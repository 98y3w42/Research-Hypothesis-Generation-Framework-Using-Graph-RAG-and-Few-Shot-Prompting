import openpyxl
import json
import os

def get_moose(dir):
    wb = openpyxl.load_workbook(dir)
    sheet = wb['Overall']
    row = 1
    column = 1
    while sheet.cell(row=row, column=1).value != None:
        row += 1
    row -= 1
    while sheet.cell(row=1, column=column).value != None:
        column += 1
    column -= 1
    chem = []
    for i in range(row):
        temp = []
        for j in range(column):
            temp.append(sheet.cell(row=i + 1, column=j + 1).value)
        chem.append(temp)
    return chem

def get_json(dir):
    temp = None
    with open(dir, 'r') as f:
        temp = json.load(f)
    return temp

def save_train(moose_json, name='./data/step0.json'):
    with open(name, 'w') as f:
        for line in moose_json:
            f.write(json.dumps(line) + "\n")

def save_json(data, filename):
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4, separators=(',', ': '))

if __name__ == '__main__':
    moose_dir = os.getcwd() + "\\data\\chem_research_2024.xlsx"
    corpus_dir = os.getcwd() + "\\data\\Inspiration_Corpus_3000.json"
    
    chems = get_moose(moose_dir)
    corpus = {paper[0].replace("\n", ""):paper[1].replace("\n", "") for paper in get_json(corpus_dir)}
    
    moose_json = []
    id = 0
    for chem in chems[1:]:
        moose = {}
        moose["id"] = id
        answers = []
        for i in range(9, 14, 2):
            title = chem[i].rstrip().replace("\n", " ")
            if not title == "NA" and title in corpus.keys():
                answers.append({"kb_id":corpus[title].replace('\n', ''), "text":corpus[title].replace('\n', '')})
        moose["answers"] = answers
        moose["question"] = chem[6].replace("\n", " ")
        moose["entities"] = answers
        moose_json.append(moose)
        id += 1
    
    save_train(moose_json)
    
    # 以下QuestionとMain Hypothesisのみ抽出
    qa_json = []
    for chem in chems[1:]:
        moose = {}
        moose["question"] = chem[6].replace("\n", " ")
        moose["hypothesis"] = chem[15].replace("\n", " ")
        qa_json.append(moose)
    save_json(qa_json[:5], "./data/moosechem_fs.json")
    save_json(qa_json[5:], "./data/moosechem_qa.json")
    