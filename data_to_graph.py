import os

import os

from langchain_community.document_loaders import PyPDFLoader
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_core.documents import Document
from langchain_core.prompts import PromptTemplate
from langchain_community.embeddings import OCIGenAIEmbeddings
from langchain_community.vectorstores import Neo4jVector
from langchain_openai import ChatOpenAI
from langchain_community.graphs import Neo4jGraph
from langchain_experimental.graph_transformers import LLMGraphTransformer
from langchain_community.chains.graph_qa.cypher import GraphCypherQAChain
from langchain.schema.runnable import Runnable
from langchain.schema.runnable.config import RunnableConfig
from langchain_community.vectorstores.utils import DistanceStrategy
from langchain.embeddings import OpenAIEmbeddings

from langchain_huggingface import HuggingFaceEmbeddings

from semanticscholar import SemanticScholar
from openai import OpenAI

import arxiv
import re
from pypdf import PdfReader
import json
import tiktoken

from typing import Optional

import openpyxl
import pprint

os.environ["NEO4J_URI"] = "URI"
os.environ["NEO4J_USERNAME"] = "neo4j"
os.environ["NEO4J_PASSWORD"] = "password"
os.environ["OPENAI_API_KEY"] = "OPENAI_API_KEY"

# embeddingモデル設定
embedding_model = HuggingFaceEmbeddings(
    model_name="intfloat/multilingual-e5-large"
)

# LLM設定
llm = ChatOpenAI(model_name="gpt-4o-mini", temperature=0)
client = OpenAI()

def get_moose(dir):
    wb = openpyxl.load_workbook(dir)
    sheet = wb['Overall']
    row = 1
    column = 1
    while sheet.cell(row=row, column=1).value != None:
        row += 1
    row -= 1
    while sheet.cell(row=1, column=column).value != None:
        column += 1
    column -= 1
    chem = []
    for i in range(row):
        temp = []
        for j in range(column):
            temp.append(sheet.cell(row=i + 1, column=j + 1).value)
        chem.append(temp)
    return chem

def get_json(dir):
    temp = None
    with open(dir, 'r') as f:
        temp = json.load(f)
    return temp 
    
def query_with_llm(query):
    messages=[{"role":"user","content":query}]
    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=messages,
        temperature=0
    )
    return response.choices[0].message.content

def create_graph(papers):
    # Neo4jへの接続情報を設定してgraphインスタンスを作成
    graph = Neo4jGraph()
    
    # DB内のグラフを削除するクエリ
    cypher = """
      MATCH (n)
      DETACH DELETE n;
    """
    
    # 既存グラフを削除して前回の内容をリセット
    #graph.query(cypher)  
    
    # llmを使いドキュメントをグラフに変換するtransformerを作成
    transformer = LLMGraphTransformer(llm=llm)
    text_splitter = RecursiveCharacterTextSplitter(chunk_size=200, chunk_overlap=40)
    for i in range(len(papers)):
        print(i)
        print(papers[i][0])
        documents = text_splitter.create_documents([papers[i][1]])
        docs = text_splitter.split_documents(documents)
        graph_documents = transformer.convert_to_graph_documents(docs)
        graph.add_graph_documents(
            graph_documents, 
            baseEntityLabel=True, 
            include_source=True
        )
    
    # ベクトルデータを含む検索用インデックス作成
    index = Neo4jVector.from_existing_graph(
        OpenAIEmbeddings(),
        node_label="entity", # 検索対象ノード
        text_node_properties=["id", "text"], # 検索対象プロパティ
        embedding_node_property="embedding", # ベクトルデータの保存先プロパティ
        index_name="vector_index", # ベクトル検索用のインデックス名
        keyword_index_name="entity_index", # 全文検索用のインデックス名
        search_type="hybrid" # 検索タイプに「ハイブリッド」を設定
    )

    # Cypherクエリ用のプロンプトテンプレート
    template = """
    Task: Generate Cypher queries to query the graph database.
    Instructions:
        Only use the relationship types and properties provided in the schema.
        Do not use any other relationship types or properties not provided.
        Schema: {schema}
    Note:
        Do not include explanations or apologies in your response.
        Do not answer questions that ask for anything other than creating Cypher statements.
        Include only the generated Cypher statements in your response.
        Question: {question}
    """ 

    # プロンプトの設定
    question_prompt = PromptTemplate(
        template=template, # プロンプトテンプレートをセット
        input_variables=["schema", "question"] # プロンプトに挿入する変数
    )

    # Cypherクエリを作成 → 実行 → 結果から回答を行うChainを作成
    qa = GraphCypherQAChain.from_llm(
        llm=llm,
        graph=graph,
        cypher_prompt=question_prompt,
        allow_dangerous_requests=True
    ) 
    
    return qa

def get_qa_bia_graph():
    template = """
    Task: Generate Cypher queries to query the graph database.
    Instructions:
        Only use the relationship types and properties provided in the schema.
        Do not use any other relationship types or properties not provided.
        Schema: {schema}
    Note:
        Do not include explanations or apologies in your response.
        Do not answer questions that ask for anything other than creating Cypher statements.
        Include only the generated Cypher statements in your response.
        Question: {question}
    """ 

    graph = Neo4jGraph()
    question_prompt = PromptTemplate(
        template=template,
        input_variables=["schema", "question"]
    )
    qa = GraphCypherQAChain.from_llm(
        llm=llm,
        graph=graph,
        cypher_prompt=question_prompt,
        allow_dangerous_requests=True
    ) 
    return qa


if __name__ == '__main__':
    moose_dir = "\\data\\chem_research_2024.xlsx"
    corpus_dir = "\\data\\Inspiration_Corpus_3000.json"
    
    #chem = get_moose(moose_dir)
    corpus = get_json(corpus_dir)
    #encoding = tiktoken.encoding_for_model("gpt-4")
    qa = create_graph(corpus)
    #qa = get_qa_bia_graph()